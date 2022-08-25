import warnings
from pathlib import Path
from typing import overload, Any

import openpyxl as opxl
from openpyxl.cell.cell import Cell

from ._core import RWFile


__all__ = ['RWExcel']


class RWExcel(RWFile):
    """
    Class for reading and writing to the Excel file as *.xlsx.

    In the class used a `openpyxl` library.

    Parameters
    ----------
    filepath: Path or path-like str
        path to the *.xlsx Excel file.
    autosave: bool, default=False
        if True save to file after any changes.

    Raises
    ------
    FilepathPatternError:
        if the filepath does not lead to the *.xlsx file.
    """

    FILE_SUFFIXES = {".xlsx"}

    def __init__(self, filepath: Path | str, autosave: bool = False):
        super().__init__(filepath)

        self._autosave = autosave
        self._hapi = opxl.open(self._fp)

        warnings.warn(
            "The RWExcel class has been removed in lib version 0.0.1.",
            DeprecationWarning, stacklevel=2
        )

    def active_sheet(self, title: str) -> None:
        """
        Change active sheet by name.

        Parameters
        ----------
        title: str
            sheet name.
        """
        self._hapi.active = self._hapi[title]

    def close(self):
        self._hapi.close()

    def save(self) -> None:
        """
        Save excel parser to the Excel file.
        """
        self._hapi.save(self._fp)

    @classmethod
    def new_empty(
            cls,
            filepath: Path | str,
            autosave: bool = False,
            first_sheet: str = 'Sheet',
    ):
        """
        Create new Excel file to the filepath with the first sheet.

        Parameters
        ----------
        filepath: Path or path-like str
            path to the *.xlsx Excel file.
        autosave: bool, default=False
            if True save to file after any changes.
        first_sheet: str, default='Sheet'
            name of the first sheet.

        Returns
        -------
        RWExcel
            new instance of the RWExcel.

        Raises
        ------
        FilepathPatternError:
            if the filepath does not lead to the *.xlsx file.
        """
        if filepath.exists():
            raise FileExistsError('Excel file is already exists')
        elif not filepath.parent.exists():
            filepath.parent.mkdir(parents=True)

        opxl.Workbook().save(filepath)
        inst = cls(filepath, autosave=autosave)
        inst.hapi.active.title = first_sheet
        inst.save()

        return inst

    @overload
    def cell(self, cell_name: str, value: Any = None) -> Cell:
        """
        Parameters
        ----------
        cell_name: str
            cell name in excel format (e.g. 'A1' or 'A1:B2').
        value: Any, default=None
            value to be set in the cell.

        Returns
        -------
        Cell
            excel cell
        """
        ...

    @overload
    def cell(self, row: int, col: int, value: Any = None) -> Cell:
        """
        Parameters
        ----------
        row: int
            row index.
        col: int
            column index.
        value: Any
            Value to be set in the cell.

        Returns
        -------
        Cell
            excel cell.
        """
        ...

    def cell(self, *coords, **kwargs) -> Cell:
        """
        Coords can be represented as a string or
        as a tuple of strings and column indices.

        The row and column indexes start from 0.

        It is not guaranteed to work correctly
        when trying to get multiple cells.

        Parameters
        ----------
        *coords
            cell coordinates.
        **kwargs
            At this point,only the 'value' key
            can be included. The others will be
            ignored.

        Returns
        -------
        Cell
            excel cell.
        """

        match coords:
            case (str() as cell_name,):
                cell = self._hapi.active[cell_name]
            case (int() as row, int() as col):
                cell = self._hapi.active.cell(row + 1, col + 1)
            case _:
                raise ValueError(f'Incorrect input: {coords}')
        if 'value' in kwargs:
            cell.value = kwargs['value']
            if self._autosave:
                self.save()
        return cell

    @property
    def hapi(self) -> opxl.Workbook:
        return self._hapi

    def __getitem__(self, *coords: str | tuple[int, int]) -> Cell:
        """
        Call .cell method with *coords.

        Parameters
        ----------
        *coords
            cell coordinates.

        Returns
        -------
        Cell
            excel cell.

        Raises
        ------
        AssertionError
            if .cell method return not a Cell instance

        See Also
        --------
        .cell: get excel cell.
        """
        match coords:
            case ((int(), int()),):
                coords = coords[0]
        res = self.cell(*coords)
        if isinstance(res, Cell):
            return res
        else:
            assert False, f'incorrect .cell return: {type(res)}'
