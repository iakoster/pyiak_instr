import re
from pathlib import Path
from typing import overload, Any

from ._rwf_utils import *

import openpyxl as opxl
from openpyxl.cell.cell import Cell


class RWExcel(object):

    FILENAME_PATTERN = re.compile('\w+.xlsx$')

    def __init__(self, filepath: Path | str, autosave: bool = False):
        filepath = if_str2path(filepath)
        check_filename(self.FILENAME_PATTERN, filepath)
        create_dir_if_not_exists(filepath)

        self._filepath = filepath
        self._autosave = autosave
        self._xcl = opxl.open(self._filepath)

    def active_sheet(self, title: str) -> None:
        self._xcl.active = self._xcl[title]

    def save(self) -> None:
        self._xcl.save(self._filepath)

    @classmethod
    def new_empty(
            cls,
            filepath: Path | str,
            autosave: bool = False,
            first_sheet: str = 'Sheet',
    ):
        filepath = if_str2path(filepath)
        check_filename(cls.FILENAME_PATTERN, filepath)
        create_dir_if_not_exists(filepath)
        if filepath.exists():
            raise FileExistsError('Excel file is already exists')

        wb = opxl.Workbook()
        wb.active.title = first_sheet
        wb.save(filepath)
        wb.close()

        return cls(filepath, autosave=autosave)

    @overload
    def cell(self, cell_name: str, value: Any = None) -> Cell:
        ...

    @overload
    def cell(self, row: int, col: int, value: Any = None) -> Cell:
        ...

    def cell(self, *coords, **kwargs) -> Cell:
        match coords:
            case (str() as cell_name,):
                cell = self._xcl.active[cell_name]
            case (int() as row, int() as col):
                cell = self._xcl.active.cell(row + 1, col + 1)
            case _:
                raise ValueError(f'Incorrect input: {coords}')
        if 'value' in kwargs:
            cell.value = kwargs['value']
            if self._autosave:
                self.save()
        return cell

    @property
    def filepath(self):
        return self._filepath

    @property
    def excel(self):
        return self._xcl

    def __getitem__(self, *coords: str | tuple[int, int]) -> Cell:
        match coords:
            case ((int(), int()),):
                coords = coords[0]
        res = self.cell(*coords)
        if isinstance(res, Cell):
            return res
        else:
            assert False, f'incorrect .cell return: {type(res)}'
